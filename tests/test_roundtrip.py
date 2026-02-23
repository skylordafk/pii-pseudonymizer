"""Integration test: multi-sheet xlsx with cross-sheet formulas, pseudonymization, and round-trip."""

import os
import tempfile

import openpyxl
import pytest

from pii_pseudonymizer.heuristics import analyze_all_columns
from pii_pseudonymizer.obfuscator import Obfuscator
from pii_pseudonymizer.reader import read_all_rows, read_xlsx, write_xlsx


@pytest.fixture
def test_workbook():
    """Create a multi-sheet xlsx with cross-sheet formula references."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        filepath = f.name

    wb = openpyxl.Workbook()

    # Sheet 1: Employees
    ws1 = wb.active
    ws1.title = "Employees"
    ws1.append(["employee_id", "first_name", "last_name", "email", "department"])
    ws1.append([1001, "Alice", "Johnson", "alice@example.com", "Engineering"])
    ws1.append([1002, "Bob", "Smith", "bob@example.com", "Marketing"])
    ws1.append([1003, "Alice", "Williams", "alice.w@example.com", "Engineering"])
    ws1.append([1004, "Diana", "Johnson", "diana@example.com", "HR"])

    # Sheet 2: Payroll (references Employees sheet)
    ws2 = wb.create_sheet("Payroll")
    ws2.append(["employee_id", "name_lookup", "ssn", "salary", "bonus"])
    ws2.append([1001, "=VLOOKUP(A2,Employees!A:C,2,FALSE)", "123-45-6789", 95000, 5000])
    ws2.append([1002, "=VLOOKUP(A3,Employees!A:C,2,FALSE)", "234-56-7890", 82000, 3000])
    ws2.append([1003, "=VLOOKUP(A4,Employees!A:C,2,FALSE)", "345-67-8901", 91000, 4500])
    ws2.append([1004, "=VLOOKUP(A5,Employees!A:C,2,FALSE)", "456-78-9012", 78000, 2000])

    # Sheet 3: Summary (references both sheets)
    ws3 = wb.create_sheet("Summary")
    ws3.append(["metric", "value"])
    ws3.append(["Total Employees", "=COUNTA(Employees!A:A)-1"])
    ws3.append(["Total Salary", "=SUM(Payroll!D:D)"])
    ws3.append(["Avg Bonus", "=AVERAGE(Payroll!E:E)"])
    ws3.append(["Dept Count", "=COUNTA(Employees!E2:E100)"])

    wb.save(filepath)
    wb.close()

    yield filepath

    if os.path.exists(filepath):
        os.unlink(filepath)


@pytest.fixture
def temp_paths():
    """Provide temporary paths for output files."""
    paths = {
        "output": tempfile.mktemp(suffix=".xlsx"),
        "key": tempfile.mktemp(suffix=".json"),
        "decoded": tempfile.mktemp(suffix=".xlsx"),
    }
    yield paths
    for p in paths.values():
        if os.path.exists(p):
            os.unlink(p)


class TestReadXlsx:
    """Test XLSX reading and metadata extraction."""

    def test_reads_all_sheets(self, test_workbook):
        metadata = read_xlsx(test_workbook)
        assert metadata["sheet_names"] == ["Employees", "Payroll", "Summary"]

    def test_sheet_metadata(self, test_workbook):
        metadata = read_xlsx(test_workbook)
        emp = metadata["sheets"]["Employees"]
        assert emp["row_count"] == 4
        assert len(emp["columns"]) == 5
        assert emp["headers"] == ["employee_id", "first_name", "last_name", "email", "department"]

    def test_column_samples(self, test_workbook):
        metadata = read_xlsx(test_workbook)
        emp = metadata["sheets"]["Employees"]
        email_col = next(c for c in emp["columns"] if c["name"] == "email")
        assert "alice@example.com" in email_col["samples"]

    def test_formula_detection(self, test_workbook):
        metadata = read_xlsx(test_workbook)
        formulas = metadata["formulas"]
        assert len(formulas["cross_sheet_refs"]) > 0
        # Payroll and Summary reference other sheets
        assert formulas["sheet_dependencies"]["Payroll"]
        assert formulas["sheet_dependencies"]["Summary"]


class TestHeuristicDetection:
    """Test heuristic analysis on realistic data."""

    def test_detects_employee_pii(self, test_workbook):
        metadata = read_xlsx(test_workbook)
        emp_cols = metadata["sheets"]["Employees"]["columns"]
        results = analyze_all_columns(emp_cols)

        by_name = {r["name"]: r for r in results}
        assert by_name["first_name"]["pii_type"] == "name"
        assert by_name["last_name"]["pii_type"] == "name"
        assert by_name["email"]["pii_type"] == "email"

    def test_detects_ssn(self, test_workbook):
        metadata = read_xlsx(test_workbook)
        pay_cols = metadata["sheets"]["Payroll"]["columns"]
        results = analyze_all_columns(pay_cols)

        by_name = {r["name"]: r for r in results}
        assert by_name["ssn"]["pii_type"] == "ssn"
        assert by_name["ssn"]["heuristic_score"] == "high"


class TestFullRoundTrip:
    """Test complete obfuscation -> decode round trip."""

    def test_multi_sheet_round_trip(self, test_workbook, temp_paths):
        passphrase = "test-multisheet-123"
        obfuscator = Obfuscator(passphrase)

        # Define what to obfuscate per sheet
        sheets_columns = {
            "Employees": [
                {"name": "first_name", "pii_type": "name"},
                {"name": "last_name", "pii_type": "name"},
                {"name": "email", "pii_type": "email"},
            ],
            "Payroll": [
                {"name": "ssn", "pii_type": "ssn"},
            ],
        }

        # Read and obfuscate
        metadata = read_xlsx(test_workbook)
        all_sheets_data = read_all_rows(test_workbook)
        obfuscated_sheets = {}

        for sname in metadata["sheet_names"]:
            headers, rows = all_sheets_data[sname]
            if sname in sheets_columns:
                obf_rows = obfuscator.obfuscate_rows(headers, rows, sheets_columns[sname])
                obfuscated_sheets[sname] = (headers, obf_rows)
            else:
                obfuscated_sheets[sname] = (headers, rows)

        # Write output
        write_xlsx(temp_paths["output"], obfuscated_sheets)
        obfuscator.save_key_file(temp_paths["key"], "test.xlsx", sheets_columns)

        # Verify determinism: "Alice" in rows 0 and 2
        _, emp_obf = obfuscated_sheets["Employees"]
        assert emp_obf[0]["first_name"] == emp_obf[2]["first_name"]

        # Verify determinism: "Johnson" in rows 0 and 3
        assert emp_obf[0]["last_name"] == emp_obf[3]["last_name"]

        # Non-sensitive columns unchanged
        assert emp_obf[0]["department"] == "Engineering"

        # Summary sheet unchanged
        _, summary_obf = obfuscated_sheets["Summary"]
        _, summary_orig = all_sheets_data["Summary"]
        assert summary_obf == summary_orig

        # Round-trip decode
        obfuscator2, key_data = Obfuscator.from_key_file(temp_paths["key"], passphrase)
        obf_sheets = read_all_rows(temp_paths["output"])

        for sname, sheet_meta in key_data["sheets"].items():
            cols = [
                {"name": n, "pii_type": info["pii_type"]}
                for n, info in sheet_meta["columns"].items()
                if info.get("obfuscated")
            ]
            if not cols or sname not in obf_sheets:
                continue

            headers, rows = obf_sheets[sname]
            decoded_rows = obfuscator2.deobfuscate_rows(headers, rows, cols)

            # Compare with originals
            _orig_headers, orig_rows = all_sheets_data[sname]
            for i in range(len(orig_rows)):
                for col in cols:
                    cname = col["name"]
                    original = str(orig_rows[i].get(cname, "")).strip()
                    decoded = str(decoded_rows[i].get(cname, "")).strip()
                    assert original == decoded, (
                        f"{sname} row {i + 1} col '{cname}': '{original}' != '{decoded}'"
                    )

    def test_output_preserves_all_sheets(self, test_workbook, temp_paths):
        """Output file should have all original sheets."""
        passphrase = "test-123"
        obfuscator = Obfuscator(passphrase)

        metadata = read_xlsx(test_workbook)
        all_sheets_data = read_all_rows(test_workbook)

        sheets_columns = {
            "Employees": [{"name": "email", "pii_type": "email"}],
        }

        obfuscated_sheets = {}
        for sname in metadata["sheet_names"]:
            headers, rows = all_sheets_data[sname]
            if sname in sheets_columns:
                obf_rows = obfuscator.obfuscate_rows(headers, rows, sheets_columns[sname])
                obfuscated_sheets[sname] = (headers, obf_rows)
            else:
                obfuscated_sheets[sname] = (headers, rows)

        write_xlsx(temp_paths["output"], obfuscated_sheets)

        out_meta = read_xlsx(temp_paths["output"])
        assert out_meta["sheet_names"] == metadata["sheet_names"]

        for sname in metadata["sheet_names"]:
            assert out_meta["sheets"][sname]["row_count"] == metadata["sheets"][sname]["row_count"]
