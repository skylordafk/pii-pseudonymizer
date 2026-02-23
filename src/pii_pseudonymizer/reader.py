"""XLSX file reading and writing with openpyxl — multi-sheet + formula detection."""

import re

import openpyxl

# Pattern to find cross-sheet references in formulas
# Matches: Sheet2!A1, 'Sheet Name'!A1:B5, 'My Sheet'!$A$1, etc.
_SHEET_REF_PATTERN = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_]+))!")


def read_xlsx(filepath):
    """
    Read an xlsx file and return structured metadata for ALL sheets.

    Returns dict with:
        filepath: str
        sheet_names: list[str]
        sheets: dict mapping sheet_name -> {
            headers: list[str],
            row_count: int,
            columns: list[dict] (index, name, samples, null_count, total_count, dtype_guess)
        }
        formulas: {
            cross_sheet_refs: list of {from_sheet, from_cell, to_sheets, formula}
            sheet_dependencies: dict mapping sheet_name -> list of sheets it references
            formula_columns: dict mapping sheet_name -> dict of col_name -> formula_count
        }
    """
    # First pass: read computed values for all sheets
    wb_data = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb_data.sheetnames

    sheets = {}
    for sname in sheet_names:
        ws = wb_data[sname]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            sheets[sname] = {
                "headers": [],
                "row_count": 0,
                "columns": [],
            }
            continue

        headers = [str(h) if h is not None else f"Column_{i + 1}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:]

        columns = []
        for col_idx, header in enumerate(headers):
            values = [row[col_idx] for row in data_rows if col_idx < len(row)]
            non_null = [v for v in values if v is not None and str(v).strip() != ""]
            null_count = len(values) - len(non_null)
            samples = [str(v) for v in non_null[:10]]
            dtype_guess = _guess_dtype(non_null[:20])

            columns.append(
                {
                    "index": col_idx,
                    "name": header,
                    "samples": samples,
                    "null_count": null_count,
                    "total_count": len(values),
                    "dtype_guess": dtype_guess,
                }
            )

        sheets[sname] = {
            "headers": headers,
            "row_count": len(data_rows),
            "columns": columns,
        }

    wb_data.close()

    # Second pass: read formulas (data_only=False) to detect cross-sheet links
    formulas = _detect_formulas(filepath, sheet_names)

    return {
        "filepath": filepath,
        "sheet_names": sheet_names,
        "sheets": sheets,
        "formulas": formulas,
    }


def _detect_formulas(filepath, sheet_names):
    """
    Scan all sheets for formulas and detect cross-sheet references.

    Opens the workbook with data_only=False (non-read-only) to access
    formula strings.
    """
    cross_sheet_refs = []
    sheet_dependencies = {s: set() for s in sheet_names}
    formula_columns = {s: {} for s in sheet_names}

    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
    except Exception:
        # If we can't open in formula mode, return empty results
        return {
            "cross_sheet_refs": [],
            "sheet_dependencies": {s: [] for s in sheet_names},
            "formula_columns": formula_columns,
        }

    sheet_names_set = set(sheet_names)

    for sname in sheet_names:
        ws = wb[sname]
        # Get headers from row 1
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value is not None else f"Column_{cell.column}")

        for row in ws.iter_rows(min_row=1):
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue

                # This cell contains a formula
                col_idx = cell.column - 1
                col_name = headers[col_idx] if col_idx < len(headers) else f"Column_{cell.column}"

                # Count formulas per column
                formula_columns[sname][col_name] = formula_columns[sname].get(col_name, 0) + 1

                # Find cross-sheet references
                referenced_sheets = set()
                for match in _SHEET_REF_PATTERN.finditer(val):
                    quoted_name = match.group(1)
                    unquoted_name = match.group(2)
                    ref_sheet = quoted_name or unquoted_name

                    # Only track references to sheets that actually exist
                    if ref_sheet in sheet_names_set and ref_sheet != sname:
                        referenced_sheets.add(ref_sheet)

                if referenced_sheets:
                    cell_ref = (
                        f"{cell.column_letter}{cell.row}"
                        if hasattr(cell, "column_letter")
                        else f"R{cell.row}C{cell.column}"
                    )
                    cross_sheet_refs.append(
                        {
                            "from_sheet": sname,
                            "from_cell": cell_ref,
                            "to_sheets": sorted(referenced_sheets),
                            "formula": val,
                        }
                    )
                    sheet_dependencies[sname].update(referenced_sheets)

    wb.close()

    return {
        "cross_sheet_refs": cross_sheet_refs,
        "sheet_dependencies": {s: sorted(deps) for s, deps in sheet_dependencies.items()},
        "formula_columns": formula_columns,
    }


def read_all_rows(filepath, sheet_name=None):
    """
    Read all data rows from one or all sheets.

    If sheet_name is given, returns (headers, rows) for that single sheet.
    If sheet_name is None, returns dict: {sheet_name: (headers, rows)} for all sheets.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if sheet_name:
        result = _read_sheet_rows(wb[sheet_name])
        wb.close()
        return result

    all_sheets = {}
    for sname in wb.sheetnames:
        all_sheets[sname] = _read_sheet_rows(wb[sname])
    wb.close()
    return all_sheets


def _read_sheet_rows(ws):
    """Read headers and data rows from a single worksheet."""
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], []

    headers = [str(h) if h is not None else f"Column_{i + 1}" for i, h in enumerate(header_row)]
    data = []
    for row in rows_iter:
        row_dict = {}
        for i, header in enumerate(headers):
            row_dict[header] = row[i] if i < len(row) else None
        data.append(row_dict)
    return headers, data


def write_xlsx(filepath, sheets_data, single_sheet=None):
    """
    Write data to a new xlsx file with one or multiple sheets.

    Args:
        filepath: output path
        sheets_data: either:
            - dict {sheet_name: (headers, rows)} for multi-sheet
            - tuple (headers, rows) when single_sheet is given
        single_sheet: if provided, treat sheets_data as a single (headers, rows)
                      tuple and write it under this sheet name
    """
    import os

    os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)

    wb = openpyxl.Workbook()

    if single_sheet:
        headers, rows = sheets_data
        ws = wb.active
        ws.title = single_sheet
        _write_sheet(ws, headers, rows)
    else:
        first = True
        for sname, (headers, rows) in sheets_data.items():
            if first:
                ws = wb.active
                ws.title = sname
                first = False
            else:
                ws = wb.create_sheet(title=sname)
            _write_sheet(ws, headers, rows)

    wb.save(filepath)
    wb.close()


def _write_sheet(ws, headers, rows):
    """Write headers and row dicts to a worksheet."""
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row_dict in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_dict.get(header)
            ws.cell(row=row_idx, column=col_idx, value=value)


def _guess_dtype(values):
    """Guess data type from sample values."""
    if not values:
        return "empty"

    numeric_count = 0
    date_count = 0
    text_count = 0

    for v in values:
        if isinstance(v, (int, float)):
            numeric_count += 1
        elif hasattr(v, "strftime"):
            date_count += 1
        else:
            s = str(v).strip()
            try:
                float(s.replace(",", ""))
                numeric_count += 1
            except (ValueError, AttributeError):
                text_count += 1

    total = len(values)
    if numeric_count / total > 0.8:
        return "numeric"
    if date_count / total > 0.8:
        return "date"
    if text_count / total > 0.8:
        return "text"
    return "mixed"
